import csv
import datetime
import uploading_data
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit
import os.path


def set_column_w(ws):
    """
    Установить одинаковую ширину для ячеек таблицы по самому длинному слову
    Args:
        ws (any): Текущий рабочий лист
    """
    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) + 2 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length


def as_text(value):
    """
    Переводит значение  в строку
    Args:
        value (str | int | float | None): Значение

    Returns:
        str: Строковое представление значения
    """
    if value is None:
        return ""
    return str(value)


ru_words = {"name": "Название",
            "description": "Описание",
            "key_skills": "Навыки",
            "experience_id": "Опыт работы",
            "premium": "Премиум-вакансия",
            "employer_name": "Компания",
            "salary": "Оклад",
            "area_name": "Название региона",
            "published_at": "Дата публикации вакансии"}
currency_to_rub = {"AZN": 35.68,
                   "BYR": 23.91,
                   "EUR": 59.90,
                   "GEL": 21.74,
                   "KGS": 0.76,
                   "KZT": 0.13,
                   "RUR": 1,
                   "UAH": 1.64,
                   "USD": 60.66,
                   "UZS": 0.0055}


class Report:
    """
    Класс для подготовки и создания отчетов
    Attributes:
        data (list): Входящие данные
    """

    def __init__(self, data):
        """
        Инициализвция класса Report
        Args:
            data (list): Входящие данные
        """
        self.data = data

    def generate_excel(self):
        """
        Подготовка и генерация exel файла в котором содержится: статистика по годам и статистика по городам
        """
        headers = ["Год", "Средняя зарплата", f"Средняя зарплата - {self.data[6]}", "Количество вакансий",
                   f"Количество вакансий - {self.data[6]}"]
        thin = Side(border_style="thin")
        wb = Workbook()
        ws0 = wb.active
        ws0.title = "Статистика по годам"
        ws1 = wb.create_sheet("Статистика по городам")

        # ws0
        for i, head in enumerate(headers):
            ws0.cell(1, i + 1, head).font = Font(bold=True)
        for year, value in self.data[0].items():
            ws0.append([year, value, self.data[2][year], self.data[1][year], self.data[3][year]])
        set_column_w(ws0)
        for column in ws0.columns:
            for cell in column:
                cell.border = Border(thin, thin, thin, thin)

        # ws1
        headers = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        s_b_c = list(self.data[4].items())
        v_b_c = list(self.data[5].items())
        for i, head in enumerate(headers):
            if i + 1 != 3:
                ws1.cell(1, i + 1, head).font = Font(bold=True)
        for i in range(10):
            ws1.append([s_b_c[i][0], s_b_c[i][1], "", v_b_c[i][0], v_b_c[i][1]])
        i = 0
        for columns in ws1["A1":"E11"]:
            for cell in columns:
                if i != 2:
                    cell.border = Border(thin, thin, thin, thin)
                if i == 4:
                    cell.number_format = FORMAT_PERCENTAGE_00
                i += 1
            i = 0
        set_column_w(ws1)

        wb.save('report.xlsx')

    def generate_image(self):
        """
        Подготовка и генерация png файла в котором содержится 4 варианта графиков:
        1) диаграмма - уровень зарплат по годам для вывода динамики уровня зарплат по годам как общий,
        так и для выбранной профессии
        2) диаграмма - количество вакансий по годам как общий, так и для выбранной профессии
        3) горизонтальная диаграмма - уровень зарплат по городам
        4) круговая диаграмма - количество вакансий по городам
        """
        plt.rcParams['font.size'] = "8"
        ratio = 1920 / 1440
        # first graph
        labels = list(self.data[0].keys())
        salary_years = [x for x in self.data[0].values()]
        salary_years_vac = [x for x in self.data[2].values()]

        x = np.arange(len(labels))
        width = 0.35

        fig, ax = plt.subplots(2, 2, figsize=(10, 10 / ratio))
        ax[0, 0].bar(x - width / 2, salary_years, width, label='средняя з/п')
        ax[0, 0].bar(x + width / 2, salary_years_vac, width, label=f'з/п {self.data[-1]}')

        ax[0, 0].set_title('Уровень зарплат по годам', fontsize=20)
        ax[0, 0].set_xticks(x, labels, rotation=90)
        ax[0, 0].legend()
        ax[0, 0].grid(axis="y")

        # second graph
        vacs_years = [x for x in self.data[1].values()]
        current_vac_years = [x for x in self.data[3].values()]
        ax[0, 1].bar(x - width / 2, vacs_years, width, label='Количество вакансий')
        ax[0, 1].bar(x + width / 2, current_vac_years, width, label=f'Количество вакансий {self.data[-1]}')

        ax[0, 1].set_title('Количество вакансий по годам', fontsize=20)
        ax[0, 1].set_xticks(x, labels, rotation=90)
        ax[0, 1].legend()
        ax[0, 1].grid(axis="y")

        # third graph
        cities = []
        for x in list(self.data[4].keys()):
            if "-" in x:
                cities.append(x.replace("-", "-\n"))
            else:
                cities.append(x.replace(" ", "\n"))
        cities.reverse()
        ax[1, 0].set_title('Уровень зарплат по городам', fontsize=20)
        ax[1, 0].barh(cities, list(self.data[4].values().__reversed__()))
        ax[1, 0].set_yticks(np.arange(len(cities)), cities, fontsize=6, horizontalalignment="right",
                            verticalalignment='center')
        ax[1, 0].grid(axis="x")

        # fourth graph
        other = 1
        proportion = 0
        for i in self.data[5].values():
            proportion += i
        other -= round(proportion, 4)
        ax[1, 1].set_title('Доля вакансий по городам', fontsize=20)
        ax[1, 1].pie([round(other, 4)] + list(list(self.data[5].values())),
                     labels=["Другие"] + list(self.data[5].keys()))
        ax[1, 1].axis('equal')

        # save to png
        plt.tight_layout(pad=1, w_pad=2, h_pad=3)
        plt.savefig("graph.png")

    def generate_pdf(self):
        """
        Подготовка и генерация pdf файла в котором содержится таблица exel и png файл
        """
        if not os.path.exists("graph.png"):
            Report.generate_image(self)
        if not os.path.exists("report.xlsx"):
            Report.generate_excel(self)
        headers1 = ["Год", "Средняя зарплата", f"Средняя зарплата - {self.data[6]}", "Количество вакансий",
                    f"Количество вакансий - {self.data[6]}"]
        headers2 = ["Город", "Уровень зарплат", "Город", "Доля вакансий"]
        image = os.path.abspath("graph.png")

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        pdf_template = template.render({'prof': self.data[6],
                                        "headers1": headers1,
                                        "headers2": headers2,
                                        "data": self.data,
                                        "image": image
                                        })

        config = pdfkit.configuration(wkhtmltopdf=r'I:\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": None})


class DataSet:
    """
    Класс для подготовки данных для дальнейшего использования.
    Attributes:
        file_name (str): Название файла
        vacancies_objects (list[Vacancy]): Список объектов Vacancy
    """

    def __init__(self, file_name):
        """
        Инициализирует объект DataSet
        Args:
            file_name (str): Название файла
        """
        self.file_name = file_name
        self.vacancies_objects = DataSet.parse_row(file_name)

    @staticmethod
    def сsv_reader(file_name):
        """
        Обрабатывает csv файл, пропуская незаполенные строки

        Args:
            file_name (str): Название файла

        Returns:
            list, list: Список названий заголовков, Список вакансий
        """
        df = pd.read_csv(file_name)
        DataSet.check_file(df)
        columns = df.columns.to_list()
        df.dropna(axis=0,how="any", inplace=True)
        result = df.values.tolist()
        return columns, result

    @staticmethod
    def check_file(list_data):
        """
        Проверяет файл на пустоту или отсутствие данных и если это подтведилось - выходит из программы

        Args:
            list_data (DataFrame): Все строки в файле
        """
        if len(list_data) == 0:
            print("Пустой файл")
            exit()
        if len(list_data) == 1:
            print("Нет данных")
            exit()

    @staticmethod
    def parse_row(file_name):
        """
        Парсит каждую профессию и преобразует в объект Vacancy
        Args:
            file_name (str): Название файла

        Returns:
            list: Возвращает список объектов Vacancy
        """
        name, rows = DataSet.сsv_reader(file_name)
        result = []
        for row in rows:
            new_row = dict(zip(name, row))
            result.append(Vacancy(new_row))
        return result


class InputConnect:
    """
    Класс для получения, обработки и печати данных необходимых пользователю
    Attributes:
        file_name (str): Название файла
        filter_dict (list[str]): Параметр фильтрации
    """

    def __init__(self):
        """
        Инициализирует класс InputConnect
        """
        self.file_name, self.filter_dict = InputConnect.get_params()

    @staticmethod
    def get_params():
        """
        Получает на вход данные введенные пользователем и возвращает их
        Returns:
            str, str: Название файла и параметр фильтрации
        """
        file_name = input("Введите название файла: ")
        # file_name = "vacancies_by_year.csv"  # использовал для генерации файла report.xlsx, graph.png
        filter_dict = input("Введите название профессии: ")
        # filter_dict = "Программист"  # использовал для генерации файла report.xlsx, graph.png
        return file_name, filter_dict

    @staticmethod
    def print_data(data, prof):
        """
        Печатает подсчитанные данные в консоль и возвращает их
        Args:
            data (list[Vacancy]): Список объектов Vacancy
            prof (str): Название профессии

        Returns:
            list: Динамика уровня зарплат по годам, Динамика количества вакансий по годам,
            Динамика уровня зарплат по годам для выбранной профессии,
            Динамика количества вакансий по годам для выбранной профессии,
            Уровень зарплат по городам (в порядке убывания),
            Доля вакансий по городам (в порядке убывания),
            Название профессии
        """
        salary_by_cities, vacs_by_cities, salary_by_years, \
        vacs_by_years, vac_salary_by_years, vac_counts_by_years = InputConnect.prepare_data(data, prof)

        print("Динамика уровня зарплат по годам:", salary_by_years)
        print("Динамика количества вакансий по годам:", vacs_by_years)
        print("Динамика уровня зарплат по годам для выбранной профессии:", vac_salary_by_years)
        print("Динамика количества вакансий по годам для выбранной профессии:", vac_counts_by_years)
        print("Уровень зарплат по городам (в порядке убывания):", salary_by_cities)
        print("Доля вакансий по городам (в порядке убывания):", vacs_by_cities)
        return [salary_by_years, vacs_by_years, vac_salary_by_years, vac_counts_by_years, salary_by_cities,
                vacs_by_cities, prof]

    @staticmethod
    def prepare_data(data,prof):
        """
        Подготовка данных к печати в консоль
        Args:
            data (list[Vacancy]): Список объектов Vacancy
            prof (str): Название профессии
        Returns:
            list[dict]: Уровень зарплат по городам,
            Доля вакансий по городам,
        """
        salary_by_years, vacs_by_years, vac_salary_by_years, vac_counts_by_years = uploading_data.main("csv_files", prof)
        area_dict = {}
        vacs_dict = {}
        for vac in data:
            InputConnect.fill_area_dict(area_dict, vac)
            InputConnect.fill_vacs_dict(vac, vacs_dict)

        salary_by_cities = InputConnect.prepare_salary_by_cities(area_dict, data)
        vacs_by_cities = InputConnect.prepare_vacs_by_cities(data, vacs_dict)
        return salary_by_cities, vacs_by_cities, salary_by_years, vacs_by_years, vac_salary_by_years, vac_counts_by_years

    # @staticmethod
    # def date_formatting_main(vac):
    #     year = int(datetime.datetime.strptime(vac.published_at, "%Y-%m-%dT%H:%M:%S%z").strftime("%Y"))
    #     return year

    @staticmethod
    def date_formatting_v1(vac):
        """
        Получает из строки год
        Args:
            vac (Vacancy): Объект Vacancy

        Returns:
            int: год в числовом представлении
        """
        year = int(vac.published_at[:4])
        return year

    # @staticmethod
    # def date_formatting_v2(vac):
    #     year = int(re.findall(r'20[0-9][0-9]',vac.published_at)[0])
    #     return year

    @staticmethod
    def fill_years(data, years):
        """
        Заполнение уникальными годами
        Args:
            data (list[Vacancy]): Список объектов Vacancy
            years (set): Уникальные года

        Returns:
            set: Уникальные года
        """
        for vac in data:
            years.add(int(datetime.datetime.strptime(vac.published_at, "%Y-%m-%dT%H:%M:%S%z").strftime("%Y")))
        years = sorted(list(years))
        years = list(range(min(years), max(years) + 1))
        return years

    @staticmethod
    def prepare_vacs_by_cities(data, vacs_dict):
        """
        Вычисляет долю вакансий по городам
        Args:
            data (list[Vacancy]): Список объектов Vacancy
            vacs_dict (dict): Словарь в котором ключом является город, а значением - количество вакансий в нём

        Returns:
            dict: Доля вакансий по городам
        """
        vacs_count = {x: round(y / len(data), 4) for x, y in vacs_dict.items()}
        vacs_count = {k: value for k, value in vacs_count.items() if value >= 0.01}
        vacs_by_cities = dict(sorted(vacs_count.items(), key=lambda item: item[1], reverse=True))
        vacs_by_cities = dict(list(vacs_by_cities.items())[:10])
        return vacs_by_cities

    @staticmethod
    def prepare_salary_by_cities(area_dict, data):
        """
        Вычисляет оклад вакансий по городам
        Args:
            area_dict (dict): Словарь в котором ключом является город, а значением - список окладов в нём
            data (list[Vacancy]): Список объектов Vacancy

        Returns:
            dict: Оклад вакансий по городам
        """
        area_list = [x for x in area_dict.items() if len(x[1]) / len(data) > 0.01]
        area_list = sorted(area_list, key=lambda item: sum(item[1]) / len(item[1]), reverse=True)
        salary_by_cities = {item[0]: int(sum(item[1]) / len(item[1])) for item in area_list[0: min(len(area_list), 10)]}
        return salary_by_cities

    @staticmethod
    def prepare_salary_by_years(salary_by_years, vac_salary_by_years):
        """
        Вычисляет динамику уровня зарплат по годам, как всех профессий, так и конкретной
        Args:
            salary_by_years (dict): Динамика количества вакансий по годам
            vac_salary_by_years (dict): Динамика уровня зарплат по годам для выбранной профессии

        Returns:
            dict, dict: Динамика количества вакансий по годам, Динамика уровня зарплат по годам для выбранной профессии
        """
        salary_by_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                           salary_by_years.items()}
        vac_salary_by_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                               vac_salary_by_years.items()}
        return salary_by_years, vac_salary_by_years

    @staticmethod
    def fill_vacs_dict(vac, vacs_dict):
        """
        Заполняет словарь количеством вакансий в городах
        Args:
            vac (Vacancy): Объект Vacancy
            vacs_dict (dict): Словарь в котором ключом является город, а значением - количество вакансий в нём
        """
        if vac.area_name in vacs_dict:
            vacs_dict[vac.area_name] += 1
        else:
            vacs_dict[vac.area_name] = 1

    @staticmethod
    def fill_area_dict(area_dict, vac):
        """
        Заполняет словарь окладом вакансий в городах
        Args:
            area_dict (dict): Словарь в котором ключом является город, а значением - список окладов в нём
            vac (Vacancy): Объект Vacancy
        """
        if vac.area_name in area_dict:
            area_dict[vac.area_name].append(vac.salary.get_salary_ru())
        else:
            area_dict[vac.area_name] = [vac.salary.get_salary_ru()]

    @staticmethod
    def create_dicts(years):
        """
        Создает пустые словари
        Args:
            years (set): Уникальные года

        Returns:
            dict, dict, dict, dict: Пустые словари salary_by_years, vac_counts_by_years,
            vac_salary_by_years, vacs_by_years
        """
        salary_by_years = {year: [] for year in years}
        vacs_by_years = {year: 0 for year in years}
        vac_salary_by_years = {year: [] for year in years}
        vac_counts_by_years = {year: 0 for year in years}
        return salary_by_years, vac_counts_by_years, vac_salary_by_years, vacs_by_years


class Vacancy:
    """
    Класс для предоставления вакансии
    Attributes:
        name (str): Назваеие профессии
        salary (object(Salary)): Класс Salary
        area_name (str): Расположение
        published_at (str): Дата публикации
    """

    def __init__(self, row):
        """
        Инициализирует объект Vacancy
        Args:
            row (dict): Словарь в котором содержиться строка вакансии, разделённая на различные пункты
        """
        self.name = row["name"]
        self.salary = Salary(row["salary_from"], row["salary_to"], row["salary_currency"])
        self.area_name = row["area_name"]
        self.published_at = row["published_at"]


class Salary:
    """
    Класс для предоставления зарплаты
    Attributes:
        salary_from (int): Нижняя граница вилки оклада
        salary_to (int): Верхняя граница вилки оклада
        salary_currency (str): Валюта оклада
        salary_ru (int): Средний оклад в рублях
    """

    def __init__(self, salary_from, salary_to, salary_currency):
        """
        Инициализирует объект Salary
        Args:
            salary_from (int): Нижняя граница вилки оклада
            salary_to (int): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.salary_ru = int((float(self.salary_from) + float(self.salary_to)) / 2) * currency_to_rub[
            self.salary_currency]

    def get_salary_ru(self):
        """
        Получение Среднего оклада в рублях
        Returns:
            int: Средний оклад в рублях
        """
        return self.salary_ru

